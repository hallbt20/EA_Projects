import pandas as pd
from re import sub
import tkinter as tk
from tkinter import ttk
from openpyxl.utils import get_column_letter
from tkinter import filedialog, messagebox
import os


def find_file(folder_path, file_start):
    for file_name in os.listdir(folder_path):
        if file_name.startswith(file_start):
            return os.path.join(folder_path, file_name)
    return None


def browse_for_file(missing_files):
    for key, value in missing_files.items():
        if not value:
            file_path = filedialog.askopenfilename(title=f"Locate {key}")
            if file_path:
                missing_files[key] = file_path
            else:
                messagebox.showwarning("File not found", f"{key} file not found. Please locate the file.")
    return missing_files


def start_window():
    root = tk.Tk()
    root.withdraw()  # Hide the main window

    folder_path = filedialog.askdirectory(title="Select a folder")
    if not folder_path:
        messagebox.showerror("Error", "No folder selected.")
        return None

    files_to_find = {
        "adv_out_file": "ADV OUT Pipeline",
        "sf_active_file": "Salesforce Active",
        "sf_closed_file": "Salesforce Closed",
        "ns_active_file": "Netsuite Active",
        "ns_closed_file": "Netsuite Closed",
        "eag_gc_oit_file": "EAG GC OIT",
        "hubspot_file": "hubspot",
        "great_lakes_file": "EAG GL",
        'originators_list': 'Originators List'
    }

    found_files = {key: find_file(folder_path, start) for key, start in files_to_find.items()}

    missing_files = {key: value for key, value in found_files.items() if value is None}

    if missing_files:
        messagebox.showinfo("Missing Files", "Some files were not found. Please locate the missing files.")
        found_files.update(browse_for_file(missing_files))

    file_path_vars = [
        found_files['adv_out_file'],
        found_files['sf_active_file'],
        found_files['sf_closed_file'],
        found_files['ns_active_file'],
        found_files['ns_closed_file'],
        found_files['eag_gc_oit_file'],
        found_files['hubspot_file'],
        found_files['great_lakes_file'],
        found_files['originators_list']
    ]

    return file_path_vars


def find_sheet_name(file_path, sheet_name_start):
    sheet_names = pd.ExcelFile(file_path).sheet_names
    for sheet_name in sheet_names:
        if sheet_name.startswith(sheet_name_start):
            return sheet_name
    return None


def dataframe_ordering(df):
    df = df[[
        'Data Source',
        'Opportunity ID',
        'Service Line Group (EA)',
        'Stage',
        'Stage (adjusted)',
        'Account Name: Account Name',
        'Opportunity Name',
        "First Year Fees (EA's portion)",
        "Total Contract Value (EA's portion)",
        'Created Date',
        'Close Date',
        'Age',
        'Opportunity Originator',
        'Opportunity Leader',
        'Opportunity Team',
        'Service Lines',
        'Type',
        'Account Name: Industry',
        'Office Location Client Assigned to',
        'Recurrence',
        'Contract Duration',
        'Last Activity',
        'Next Step',
        'Next Step Due Date',
        'Client Code',
        'Primary Campaign Source: Campaign Name',
        'Originator Service Line',
        'Opp Leader Service Line',
        'Contact'
    ]]

    return df


def excel_formatter(list_of_df, file_location, list_of_sheet_names):
    """
    Save DataFrames to an Excel file with auto-fitted columns, accessible filters, and expanded first row height.

    Parameters:
    list_of_df (list of pd.DataFrame): The DataFrames to be written to the Excel file.
    file_location (str): The path to the Excel file to be saved.
    list_of_sheet_names (list of str): The names of the sheets to be created in the Excel file.
    """
    # Create a Pandas Excel writer using openpyxl as the engine
    with pd.ExcelWriter(file_location, engine='openpyxl') as writer:
        for index, df in enumerate(list_of_df):
            # Write the DataFrame to the Excel file
            df.to_excel(writer, sheet_name=list_of_sheet_names[index], index=False)

            # Access the workbook and the sheet
            workbook = writer.book
            worksheet = writer.sheets[list_of_sheet_names[index]]

            # Auto-fit columns based on the widest cell in each column
            for column_cells in worksheet.columns:
                max_length = 0
                column = column_cells[0].column_letter  # Get the column name
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 4)  # Increased padding
                worksheet.column_dimensions[column].width = adjusted_width

            # Expand the height of the first row
            worksheet.row_dimensions[1].height = 30

            # Add filters
            worksheet.auto_filter.ref = worksheet.dimensions

    # Save the workbook
    workbook.save(file_location)


def clean_legacy(df):
    cols_to_initialize = [
        'Office Location Client Assigned to', 'Recurrence', 'Contract Duration', 'Last Activity',
        'Next Step', 'Next Step Due Date', 'Client Code', 'Primary Campaign Source: Campaign Name',
        'Originator Service Line', 'Opp Leader Service Line', 'Contact'
    ]

    for col in cols_to_initialize:
        df[col] = None

    # Manually renaming the columns
    df = df.rename(columns={
        "Data Source": "Data Source",
        "Record ID": "Opportunity ID",
        "Service Line Group": "Service Line Group (EA)",
        "STAGE": "Stage",
        "Stage Adjusted": "Stage (adjusted)",
        "ACCOUNT_NAME": "Account Name: Account Name",
        "OPPORTUNITY_NAME": "Opportunity Name",
        "ESTIMATED_FEES": "First Year Fees (EA's portion)",
        "Total Contract Value": "Total Contract Value (EA's portion)",
        "CREATED_DATE": "Created Date",
        "EXPECTED_CLOSE": "Close Date",
        "ORIGINATOR": "Opportunity Originator",
        "SALES_LEADER": "Opportunity Leader",
        "Team": "Opportunity Team",
        "SERVICE_LINE": "Service Lines",
        "TYPE": "Type",
        "INDUSTRY": "Account Name: Industry",
    })

    return dataframe_ordering(df)


def clean_triangle(df, active=True):
    cols_to_initialize = [
        'Recurrence',
        'Contract Duration',
        'Last Activity',
        'Next Step',
        'Next Step Due Date',
        'Client Code',
        'Primary Campaign Source: Campaign Name',
        'Originator Service Line',
        'Opp Leader Service Line',
        'Contact'
    ]

    for col in cols_to_initialize:
        df[col] = None

    df.rename(columns={
        "Data source": "Data Source",
        "Opp ID": "Opportunity ID",
        "Service Line Group": "Service Line Group (EA)",
        "Stage adjusted": "Stage (adjusted)",
        "Account Name": "Account Name: Account Name",
        "Account name": "Account Name: Account Name",
        "Opp Name": "Opportunity Name",
        "Closed Date": "Close Date",
        "First Year Fees": "First Year Fees (EA's portion)",
        "Total Contract Value": "Total Contract Value (EA's portion)",
        "Originator": "Opportunity Originator",
        "Leader": "Opportunity Leader",
        "Team": "Opportunity Team",
        "Service Line": "Service Lines",
        "Industry/Segment": "Account Name: Industry",
        "Industry Group": "Account Name: Industry",
        "Office Location": "Office Location Client Assigned to"
    }, inplace=True)

    return dataframe_ordering(df)


def clean_salesforce(df):
    """
    Function that cleans and formats both active and closed Salesforce data
    :param df: Dataframe containing active or closed records from Salesforce
    :return: Updated dataframe with added columns, duplicates removed, and proper formatting
    """
    # Add columns to dataframe
    df['Data Source'] = 'Salesforce'
    df['Stage (adjusted)'] = df['Stage']

    # Initialize empty columns
    df['Originator Service Line'] = None
    df['Opp Leader Service Line'] = None
    df['Contact'] = None

    # Remove duplicates and keep only the first occurrence for 'Opportunity ID' (this is case-sensitive)
    df = df.drop_duplicates(subset='Opportunity ID', keep='first')

    return dataframe_ordering(df)


def clean_netsuite(df, active=True):
    # Add columns 'Data source' to dataframes
    df['Data Source'] = 'NetSuite'
    df['Contract Duration'] = None
    df['Last Activity'] = None
    df['Next Step'] = None
    df['Next Step Due Date'] = None
    df['Client Code'] = None
    df['Primary Campaign Source: Campaign Name'] = None

    # Fill empty 'Service Description' values with corresponding value from 'Service'
    df['Service Description'] = df['Service Description'].fillna(df['Service'])

    # Create 'Stage (adjusted)' column based on 'Service Status' values
    if active:
        df['Stage (adjusted)'] = df['Service Status'].apply(
            lambda x: 'Qualified' if x == 'Prospect - Active' else 'Proposal'
        )
    else:
        df['Stage (adjusted)'] = df['Service Status'].apply(
            lambda x: 'Closed Won' if x == 'Won' else 'Closed Lost'
        )

    df['Total Contract Value'] = df['Estimated Fee']

    df.rename(columns={
        'Opp Number': 'Opportunity ID',
        'Service Line L1': 'Service Line Group (EA)',
        'Service Status': 'Stage',
        'Organization': 'Account Name: Account Name',
        'Service Description': 'Opportunity Name',
        'Estimated Fee': 'First Year Fees (EA\'s portion)',
        'Total Contract Value': 'Total Contract Value (EA\'s portion)',
        'Create Date': 'Created Date',
        'Service Status Change Date': 'Close Date',
        'Days Open': 'Age',
        'Originator': 'Opportunity Originator',
        'Opp Leader': 'Opportunity Leader',
        'Other Contributors': 'Opportunity Team',
        'Service': 'Service Lines',
        'Opportunity Type': 'Type',
        'Industry': 'Account Name: Industry',
        'Office': 'Office Location Client Assigned to',
        'Recurring or One Time?': 'Recurrence'
    }, inplace=True)

    return dataframe_ordering(df)


def clean_pnt(df, active=True):
    df = df.copy()
    df['Data Source'] = 'PNT Connectwise'
    df['Service Line Group (EA)'] = 'OUT'

    cols_to_initialize = [
        'Age', 'Opportunity Team', 'Account Name: Industry', 'Office Location Client Assigned to', 'Recurrence',
        'Contract Duration', 'Last Activity', 'Next Step', 'Next Step Due Date',
        'Client Code', 'Primary Campaign Source: Campaign Name', 'Originator Service Line',
        'Opp Leader Service Line', 'Contact'
    ]

    for col in cols_to_initialize:
        df[col] = None

    if active:
        df['Stage (adjusted)'] = df['Sales_Stage']
    else:
        df['Stage (adjusted)'] = df['Sales_Stage'].apply(lambda x: 'Closed Won' if 'Won' else 'Closed Lost')

    df['Total Contract Value (EA\'s portion)'] = df['Total_Revenue'] - df['Product_Cost']
    df.drop(columns=['Total_Revenue', 'Product_Cost'], inplace=True)

    df['First Year Fees (EA\'s portion)'] = df['Total Contract Value (EA\'s portion)']

    df.rename(columns={
        'RecID': 'Opportunity ID',
        'Sales_Stage': 'Stage',
        'Company_Name': 'Account Name: Account Name',
        'Opp_Name': 'Opportunity Name',
        'Created_Date': 'Created Date',
        'Expected_Close_Date': 'Close Date',
        'Originator': 'Opportunity Originator',
        'Opp_Leader': 'Opportunity Leader',
        'Service_Area': 'Service Lines',
        'Service_Type': 'Type',
    }, inplace=True)

    return dataframe_ordering(df)


def clean_hubspot(df):
    df['Data Source'] = 'Hubspot'
    df['Service Line Group (EA)'] = 'OUT'

    cols_to_initialize = [
        'Created Date', 'Age', 'Opportunity Originator', 'Service Lines', 'Office Location Client Assigned to',
        'Recurrence', 'Contract Duration', 'Last Activity', 'Next Step', 'Next Step Due Date', 'Client Code',
        'Primary Campaign Source: Campaign Name', 'Originator Service Line', 'Opp Leader Service Line', 'Contact'
    ]

    for col in cols_to_initialize:
        df[col] = None

    # Filter dataframe where Deal Stage is only one of the following
    df = df[df['Deal Stage'].isin([
        'Inquiry',
        'Intro Call Scheduled',
        'BD Action Required',
        'Consideration / Materials Sent',
        'EA Incoming Leads',
        'Assessment',
        'Engagement Letter Sent',
        'Dead Leads Inquiries',
        'Closed lost'
    ])]

    df = df[df['EA Opportunity'].isna()].drop(columns=['EA Opportunity'])

    df = df[df['Service Team'] == 'Startup']

    df['Stage (adjusted)'] = df['Deal Stage'].apply(
        lambda x: 'Closed Lost' if x in ['Dead Leads Inquiries', 'Closed lost'] else
        'Proposal' if x == 'Engagement Letter Sent' else
        'Qualified' if x in ['Consideration / Materials Sent', 'BD Action Required'] else
        'Unqualified' if x in ['Inquiry', 'Intro Call Scheduled', 'EA Incoming Leads', 'Assessment'] else
        'None'
    )

    df = df[~(df['Deal owner'] == 'John Delalio')]

    df.drop(columns=['Type'], inplace=True)

    df.rename(columns={
        'Record ID': 'Opportunity ID',
        'Deal Stage': 'Stage',
        'Deal Name': 'Account Name: Account Name',
        'Service Team': 'Opportunity Name',
        'Amount': 'First Year Fees (EA\'s portion)',
        'Amount in company currency': 'Total Contract Value (EA\'s portion)',
        'Deal owner': 'Opportunity Leader',
        'Source/Referral': 'Opportunity Team',
        'Deal Type': 'Type',
        'Pipeline': 'Account Name: Industry'
    }, inplace=True)

    return dataframe_ordering(df)


def clean_great_lakes_active_cols(ws):
    """
    :param ws: Worksheet for Great Lakes data
    :return: A list of the unhidden columns from the worksheet
    """
    # Find hidden columns to exclude
    max_col = ws.max_column
    cols = [get_column_letter(i) for i in range(1, max_col + 1)]

    # Find hidden columns
    hidden_cols = []
    last_hidden = 0
    for i, col in enumerate(cols):
        # Column is hidden
        if ws.column_dimensions[col].hidden:
            hidden_cols.append(col)
            # Last column in the hidden group
            last_hidden = ws.column_dimensions[col].max
        # Appending column if more columns in the group
        elif i + 1 <= last_hidden:
            hidden_cols.append(col)

    hidden_col_names = [ws[f'{col}1'].value for col in hidden_cols]

    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(row)

    df = pd.DataFrame(data[1:], columns=data[0])

    df = df.drop(columns=hidden_col_names)

    return df


def clean_great_lakes(df, active=True):
    # Columns to add
    df['Data Source'] = 'Great Lakes'
    df['Service Line Group (EA)'] = 'ADV'
    df['Stage'] = df['Stage (adjusted)']
    df['Opportunity Leader'] = 'Marc Moskowitz'
    df['Service Lines'] = 'ADV Workday Adaptive Planning'

    # Columns to initialize
    cols_to_initialize = [
        'Opportunity ID', 'Opportunity Team', 'Contract Duration', 'Last Activity', 'Next Step', 'Next Step Due Date',
        'Client Code', 'Primary Campaign Source: Campaign Name', 'Originator Service Line', 'Opp Leader Service Line',
        'Contact'
    ]

    for col in cols_to_initialize:
        df[col] = None

    df['Opportunity Name'] = df['Opportunity Name'].apply(
        lambda x: 'Workday Adaptive Planning - Implementation' if str(x).replace(' ', '').startswith('W-I') else
        'Workday Adaptive Planning - Optimization' if str(x).replace(' ', '').startswith('W-O') else
        'Workday Adaptive Planning - Add On' if str(x).replace(' ', '').startswith('W-AddOn') else
        'Workday Adaptive Planning - Prepaid Block' if str(x).replace(' ', '').startswith('W-B') else
        'Workday Adaptive Planning - TM' if str(x).replace(' ', '').startswith('W-TM') else
        'Workday Adaptive Planning - Renewal' if str(x).replace(' ', '').startswith('W-Renewal') else
        'Workday Adaptive Planning - Advisory' if str(x).startswith('Advisory') or str(x).endswith('Advisory') else
        'Workday Adaptive Planning - MS' if str(x).replace(' ', '').startswith('W-MS') else
        None
    )

    # Populate empty values in 'First Year Fees' with corresponding values from 'Total Contract Value'
    df['First Year Fees (EA\'s portion)'] = df['First Year Fees (EA\'s portion)'].fillna(
        df['Total Contract Value (EA\'s portion)'])

    if active:
        df = df[~df['Stage (adjusted)'].isin(['Closed Lost', 'Closed Won', '']) & ~df['Stage (adjusted)'].isna()]

        df['Stage (adjusted)'] = df['Stage'].apply(
            lambda x: 'Qualified' if x in ['AOTP', 'Discovery', 'Scoping', 'SQL'] else
            'Unqualified' if x in ['Cold SQL', 'MQL'] else
            'Proposal' if x in ['Renewals', 'SOW', 'SOW - No Decision', 'Verbal Approval'] else
            'None'
        )
    else:
        df = df[~df['Close Date'].isna() & (df['Close Date'] >= pd.Timestamp('2023-08-01'))]

    df = df[df['Opportunity Originator'] != 'Old Old']

    df.rename(columns={
        'Deal Stage': 'Stage',
        'Deal Name': 'Account Name: Account Name',
        'Service Team': 'Opportunity Name',
        'Amount': 'First Year Fees (EA\'s portion)',
        'Amount in company currency': 'Total Contract Value (EA\'s portion)',
        'Deal owner': 'Opportunity Leader',
        'Source/Referral': 'Opportunity Team',
        'Deal Type': 'Type',
        'Pipeline': 'Account Name: Industry'
    }, inplace=True)

    return dataframe_ordering(df)


def clean_combined(df):
    df['Type'] = df.apply(
        lambda row: 'Renewal Business' if row['Type'] in [
            'Existing Business', 'Renewal Business', 'Renewal of Existing Business'
        ]
        else 'Expanded Business' if row['Type'] in [
            'Expanded Business', 'Expansion of Existing Services', 'New Service for Existing Client',
            'New Service(s) for Existing Client'
        ]
        else 'New Business' if row['Type'] in ['NEW', 'New', 'New Business', 'New Client']
        else 'Renewal Business' if row['Type'] == 'TBD' and row['Service Lines'] == 'Maintenance Renewal'
        else 'Expanded Business' if row['Type'] == 'TBD' and row['Service Lines'] != 'Maintenance Renewal'
        else row['Type'],
        axis=1
    )

    # Fill empty 'Opportunity Originator' values with corresponding value from 'Opportunity Leader'
    df['Opportunity Originator'] = df['Opportunity Originator'].fillna(df['Opportunity Leader'])

    df.rename(columns={
        'Service Line Group (EA)': 'Service Line Group',
        'Account Name: Account Name': 'Account Name',
        'First Year Fees (EA\'s portion)': 'First Year Fees',
        'Total Contract Value (EA\'s portion)': 'Total Contract Value',
        'Account Name: Industry': 'Industry',
        'Office Location Client Assigned to': 'Office Location'
    }, inplace=True)

    return df


def update_adv_closed(df, originator_list_file, updated=False):
    originator_list = pd.read_excel(originator_list_file)

    # Apply the cleaning function to the 'Opportunity Originator' column
    df.loc[:, 'Opportunity Originator'] = df.loc[:, 'Opportunity Originator'].apply(
        lambda x: sub(r'\s*\(.*?\)$', '', x)
    )

    # Merge the two dataframes on the 'Opportunity Originator' column
    merged_df = pd.merge(
        df,
        originator_list[['Opportunity Originator', 'ADV?']],
        on='Opportunity Originator',
        how='left'
    )

    # Extract unique 'Opportunity Originator' values from unmatched records
    unique_unmatched_originators = merged_df[merged_df['ADV?'].isnull()]['Opportunity Originator'].unique()

    if updated or not len(unique_unmatched_originators):
        merged_df.rename(columns={
            'ADV?': 'Originator in ADV?'
        }, inplace=True)

        # Get the list of columns
        cols = list(merged_df.columns)

        # Move 'Originator in ADV?' behind 'Opportunity Originator'
        originator_col = cols.pop(cols.index('Originator in ADV?'))
        cols.insert(cols.index('Opportunity Originator') + 1, originator_col)

        return merged_df[cols]

    # Prompt user to determine the 'ADV?' value for each unique unmatched originator
    new_entries = prompt_adv_values(unique_unmatched_originators)

    for key in new_entries.keys():
        originator_list.loc[-1] = [key, None, new_entries[key]]
        originator_list = originator_list.reset_index(drop=True)

    originator_list = originator_list.sort_values(by='Opportunity Originator', ascending=True)

    try:
        excel_formatter([originator_list], originator_list_file, 'Sheet1')
    except PermissionError:
        root = tk.Tk()
        root.withdraw()  # Hide the root window
        messagebox.showerror("Permission Error",
                             "The file 'Originators List.xlsx' appears to be open. Please close the file and press OK.")
        root.destroy()
        excel_formatter(originator_list, originator_list_file, 'Sheet1')

    return update_adv_closed(df, True)


# Function to prompt user with Tkinter
def prompt_adv_values(unique_originators):
    root = tk.Tk()
    root.title("Fill in ADV? values")

    entries = []
    new_entries = {}

    def check_values():
        all_selected = all(entry[1].get() != "Select" for entry in entries)
        submit_button.config(state=tk.NORMAL if all_selected else tk.DISABLED)

    def on_submit():
        for entry in entries:
            originator = entry[0]
            adv_value = entry[1].get()
            new_entries[originator] = adv_value
        root.destroy()

    # Create labels and dropdowns for each unique unmatched originator
    for i, originator in enumerate(unique_originators):
        label = tk.Label(root, text=f"{originator}")
        label.grid(row=i, column=0)
        var = tk.StringVar(root)
        var.set("Select")
        dropdown = ttk.Combobox(root, textvariable=var, values=["ADV", "Other"])
        dropdown.grid(row=i, column=1)
        dropdown.bind("<<ComboboxSelected>>", lambda event: check_values())
        entries.append((originator, var))

    # Create a submit button, initially disabled
    submit_button = tk.Button(root, text="Submit", command=on_submit, state=tk.DISABLED)
    submit_button.grid(row=len(unique_originators), column=0, columnspan=2)

    # Run the Tkinter main loop
    root.mainloop()

    return new_entries


if __name__ == "__main__":
    # Starts the UI window for selecting a folder with contained files
    found_files = start_window()

    adv_out_file = found_files[0]
    legacy_oit_active_sheet = find_sheet_name(adv_out_file, 'Legacy OIT Active')
    triangle_active_sheet = find_sheet_name(adv_out_file, 'Triangle Active')
    triangle_closed_sheet = find_sheet_name(adv_out_file, 'Triangle Wins')

    # Read Legacy OIT and Triangle data as Pandas dataframes
    legacy_oit_active = clean_legacy(pd.read_excel(adv_out_file, sheet_name=legacy_oit_active_sheet))
    triangle_active = clean_triangle(pd.read_excel(adv_out_file, sheet_name=triangle_active_sheet))
    triangle_closed = clean_triangle(pd.read_excel(adv_out_file, sheet_name=triangle_closed_sheet))

    # Read, cleans, and reformats Salesforce data to a Pandas dataframe
    sf_active = clean_salesforce(pd.read_excel(found_files[1]))
    sf_closed = clean_salesforce(pd.read_excel(found_files[2]))

    # Read, clean, and NetSuite data as Pandas dataframes and returns cleaned data frame
    ns_active = clean_netsuite(pd.read_excel(found_files[3]))
    ns_closed = clean_netsuite(pd.read_excel(found_files[4]), False)

    # Read Gulf Coast Outsourced IT Data
    eag_gc_oit = pd.read_excel(found_files[5])

    pnt_cw_active = clean_pnt(eag_gc_oit[eag_gc_oit['Closed_Status'].isna()])
    pnt_cw_closed = clean_pnt(
        eag_gc_oit[
            (~eag_gc_oit['Closed_Status'].isna()) &
            (~eag_gc_oit['Closed_Date'].isna()) &
            (eag_gc_oit['Closed_Date'] >= pd.Timestamp('2023-08-01')) &
            (eag_gc_oit['Closed_Date'] <= pd.Timestamp('now'))
            ], False
    )

    hubspot_df = clean_hubspot(pd.read_excel(found_files[6]))
    hubspot_active = hubspot_df[~(hubspot_df['Stage (adjusted)'] == 'Closed Lost')]
    hubspot_closed = hubspot_df[hubspot_df['Stage (adjusted)'] == 'Closed Lost']

    great_lakes_file = found_files[7]
    great_lakes_active_sheet = find_sheet_name(great_lakes_file, 'ADV Active')
    great_lakes_closed_sheet = find_sheet_name(great_lakes_file, 'ADV Closed')

    # Read Legacy OIT and Triangle data as Pandas dataframes
    great_lakes_active = clean_great_lakes(pd.read_excel(great_lakes_file, sheet_name=great_lakes_active_sheet))
    great_lakes_closed = clean_great_lakes(pd.read_excel(great_lakes_file, sheet_name=great_lakes_closed_sheet), False)

    all_active = pd.concat([
        sf_active,
        ns_active,
        legacy_oit_active,
        triangle_active,
        pnt_cw_active,
        hubspot_active,
        great_lakes_active
    ], ignore_index=True)

    all_closed = pd.concat([
        triangle_closed,
        sf_closed,
        ns_closed,
        pnt_cw_closed,
        hubspot_closed,
        great_lakes_closed
    ], ignore_index=True)

    all_active = clean_combined(all_active)
    all_closed = clean_combined(all_closed)

    # Split off Advisory data
    adv_active = all_active[all_active['Service Line Group'].isin(['ADV', 'ADV (Advisory)'])]
    adv_closed = all_closed[all_closed['Service Line Group'].isin(['ADV', 'ADV (Advisory)'])]

    adv_closed = update_adv_closed(adv_closed, found_files[8])

    out_active = all_active[all_active['Service Line Group'].isin(['OUT', 'BOutsourced Services'])]
    out_closed = all_closed[all_closed['Service Line Group'].isin(['OUT', 'BOutsourced Services'])]

    # Create an ExcelWriter object and specify the file path
    with pd.ExcelWriter('ADV Pipeline Test.xlsx', engine='xlsxwriter') as writer:
        # Write each dataframe to a different sheet
        adv_active.to_excel(writer, sheet_name='ADV Active', index=False)
        adv_closed.to_excel(writer, sheet_name='ADV Closed', index=False)
